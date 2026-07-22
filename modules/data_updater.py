# -*- coding: utf-8 -*-
"""
data_updater.py
홍보담당자가 업로드한 마스터 xlsx를 파싱해 앱 데이터(CSV/JSON) 대비 델타를 계산하고
반영본을 생성하는 순수 로직 모듈. (streamlit 비의존 — 단독 테스트 가능)

지원 파일:
  - 언론대응이력 : ★언론대응내역_Master.xlsx  (시트 '언론대응이력(YYYY)')
  - 기자리스트   : ★출입기자 리스트_Master.xlsx (시트 '언론사_리스트')

핵심 원칙:
  - 대응이력: (발생일, 이슈보고 앞 30자) 지문으로 중복 제거 → 누락 backfill 포함 신규만 추가
  - 기자리스트: 정규 기자(이름/직급/Mobile/E-MAIL)만 diff. DESK·편집국장 엔트리는 보존(오삭제 방지)
"""
from __future__ import annotations

import datetime
import io
import json
import re
from collections import Counter

import openpyxl
import pandas as pd

HISTORY_COLS = ["순번", "발생 일시", "발생 유형", "현업 부서", "단계", "이슈 발생 보고", "대응 결과"]


# ───────────────────────── 공통 유틸 ─────────────────────────
def _clean(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).replace("\r", " ").replace("\n", " ")).strip()


def _norm_name(s) -> str:
    return re.sub(r"\s+", "", str(s or ""))


def _norm_phone(p) -> str:
    return re.sub(r"\D", "", p or "")


def _infer_type(bogo: str) -> str:
    b = bogo or ""
    if b.strip().startswith("<이슈 발생 보고>"):
        return "기자문의"
    if re.match(r"^\s*\[", b):
        return "기사게재"
    return ""


def _load_wb(xlsx_bytes: bytes):
    return openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)


# ───────────────────────── 파일 유형 감지 ─────────────────────────
def detect_file_type(xlsx_bytes: bytes) -> str | None:
    """업로드 파일이 어떤 마스터인지 시트명으로 자동 판별. 'history' | 'journalist' | None"""
    try:
        wb = _load_wb(xlsx_bytes)
        names = wb.sheetnames
        wb.close()
    except Exception:
        return None
    if any("언론사_리스트" in n or "언론사 리스트" in n for n in names):
        return "journalist"
    if any(re.search(r"언론대응이력\(\d{4}\)", n) for n in names):
        return "history"
    return None


def _latest_history_sheet(wb) -> str | None:
    yrs = []
    for n in wb.sheetnames:
        m = re.search(r"언론대응이력\((\d{4})\)", n)
        if m:
            yrs.append((int(m.group(1)), n))
    if not yrs:
        return None
    return max(yrs)[1]


# ───────────────────────── 대응이력 diff ─────────────────────────
def parse_history_delta(xlsx_bytes: bytes, current_csv: pd.DataFrame) -> dict:
    """
    업로드 xlsx의 최신연도 시트 vs 현재 CSV 비교 → 신규행 계산.
    반환: {sheet, new_rows(list[dict]), start_seq, end_seq, updated_df}
    """
    wb = _load_wb(xlsx_bytes)
    sheet = _latest_history_sheet(wb)
    if not sheet:
        wb.close()
        raise ValueError("언론대응이력(YYYY) 시트를 찾을 수 없습니다.")

    # 현재 CSV 지문
    csv = current_csv.copy()
    csv_dt = pd.to_datetime(csv["발생 일시"], errors="coerce")

    def fp(dt, title):
        return (str(dt)[:10], _clean(title)[:30])

    csv_fps = set(fp(csv_dt[i], csv.loc[i, "이슈 발생 보고"]) for i in csv.index)
    seq_vals = [int(float(x)) for x in csv["순번"] if str(x).strip() and str(x) != "nan"]
    max_num = max(seq_vals) if seq_vals else 0

    ws = wb[sheet]
    new_rows = []
    for r in list(ws.iter_rows(values_only=True))[2:]:
        if r[0] is None or str(r[0]).strip() == "":
            continue
        if not isinstance(r[1], datetime.datetime):
            continue
        bogo = _clean(r[7])
        if fp(r[1], bogo) in csv_fps:
            continue
        typ = _clean(r[2]) or _infer_type(bogo)
        new_rows.append({
            "발생 일시": r[1].strftime("%Y-%m-%d"),
            "발생 유형": typ,
            "현업 부서": _clean(r[5]),
            "단계": _clean(r[6]) or "관심",
            "이슈 발생 보고": bogo,
            "대응 결과": _clean(r[8]),
        })
    wb.close()

    new_rows.sort(key=lambda x: x["발생 일시"])
    for i, h in enumerate(new_rows):
        h["순번"] = str(max_num + 1 + i)

    add_df = pd.DataFrame([{c: h.get(c, "") for c in HISTORY_COLS} for h in new_rows])
    updated_df = pd.concat([csv[HISTORY_COLS], add_df], ignore_index=True) if new_rows else csv[HISTORY_COLS].copy()

    return {
        "sheet": sheet,
        "new_rows": new_rows,
        "start_seq": max_num + 1 if new_rows else None,
        "end_seq": max_num + len(new_rows) if new_rows else None,
        "updated_df": updated_df,
    }


def history_updated_bytes(updated_df: pd.DataFrame) -> bytes:
    return updated_df.to_csv(index=False).encode("utf-8-sig")


# ───────────────────────── 기자리스트 diff ─────────────────────────
def _parse_journalist_sheet(wb) -> dict:
    """언론사_리스트 시트 → {norm_name: {name, 구분, 담당자, reporters:[{이름,직책,연락처,이메일}]}}"""
    ws = wb["언론사_리스트"] if "언론사_리스트" in wb.sheetnames else wb[
        next(n for n in wb.sheetnames if "언론사" in n)
    ]
    rows = list(ws.iter_rows(values_only=True))
    media, cur = {}, None
    for r in rows[9:]:
        name = _clean(r[2]) if len(r) > 2 else ""
        jname = _clean(r[3]) if len(r) > 3 else ""
        if name:
            cur = _norm_name(name)
            media[cur] = {"name": name, "구분": _clean(r[0]), "담당자": _clean(r[10]) if len(r) > 10 else "", "reporters": []}
        if cur and jname:
            media[cur]["reporters"].append({
                "이름": jname,
                "직책": _clean(r[4]) if len(r) > 4 else "",
                "연락처": _clean(r[5]) if len(r) > 5 else "",
                "이메일": _clean(r[6]) if len(r) > 6 else "",
            })
    return media


def _is_desk(rep: dict) -> bool:
    return ("DESK" in str(rep.get("비고", ""))
            or rep.get("직책") in ("팀장/차장급", "편집국장")
            or "편집국장" in rep.get("이름", ""))


def parse_journalist_delta(xlsx_bytes: bytes, master: dict) -> dict:
    """
    업로드 xlsx vs master_data.json 비교 → 정규 기자 add/update/remove + 신규매체.
    DESK/편집국장 엔트리는 보존. 반환: {updated_master, summary}
    """
    import copy
    md = copy.deepcopy(master)
    mc = md.setdefault("media_contacts", {})
    mc_by_norm = {_norm_name(k): k for k in mc}

    wb = _load_wb(xlsx_bytes)
    xl_media = _parse_journalist_sheet(wb)
    wb.close()

    summary = {"add_rep": [], "upd_rep": [], "del_rep": [], "new_media": [], "unmatched_master": []}

    for nn, xm in xl_media.items():
        if nn not in mc_by_norm:
            summary["new_media"].append(xm["name"])
            mc[xm["name"]] = {
                "구분": xm["구분"], "담당자": xm["담당자"],
                "출입기자": [dict(이름=r["이름"], 직책=r["직책"], 연락처=r["연락처"], 이메일=r["이메일"]) for r in xm["reporters"]],
            }
            continue
        mkey = mc_by_norm[nn]
        entry = mc[mkey]
        reps = entry.get("출입기자", [])
        desk = [r for r in reps if isinstance(r, dict) and _is_desk(r)]
        reg = [r for r in reps if isinstance(r, dict) and not _is_desk(r)]
        reg_by_name = {r.get("이름", ""): r for r in reg}
        xl_by_name = {r["이름"]: r for r in xm["reporters"]}

        new_reg = []
        for jn, xr in xl_by_name.items():
            if jn in reg_by_name:
                mr = dict(reg_by_name[jn])
                ch = []
                if _norm_phone(xr["연락처"]) != _norm_phone(mr.get("연락처", "")):
                    ch.append(f"연락처 {mr.get('연락처', '') or '-'}→{xr['연락처'] or '-'}")
                    mr["연락처"] = xr["연락처"]
                if xr["이메일"] and xr["이메일"] != mr.get("이메일", ""):
                    ch.append(f"이메일 {mr.get('이메일', '') or '-'}→{xr['이메일']}")
                    mr["이메일"] = xr["이메일"]
                if xr["직책"] and xr["직책"] != mr.get("직책", ""):
                    mr["직책"] = xr["직책"]
                if ch:
                    summary["upd_rep"].append((mkey, jn, ch))
                new_reg.append(mr)
            else:
                summary["add_rep"].append((mkey, jn, xr["직책"], xr["연락처"]))
                new_reg.append(dict(이름=jn, 직책=xr["직책"], 연락처=xr["연락처"], 이메일=xr["이메일"]))
        for jn in reg_by_name:
            if jn not in xl_by_name:
                summary["del_rep"].append((mkey, jn))
        entry["출입기자"] = desk + new_reg

    for nn, mkey in mc_by_norm.items():
        if nn not in xl_media:
            summary["unmatched_master"].append(mkey)

    return {"updated_master": md, "summary": summary}


def journalist_updated_bytes(updated_master: dict) -> bytes:
    return json.dumps(updated_master, ensure_ascii=False, indent=2).encode("utf-8")


def has_journalist_changes(summary: dict) -> bool:
    return bool(summary["add_rep"] or summary["upd_rep"] or summary["del_rep"] or summary["new_media"])
